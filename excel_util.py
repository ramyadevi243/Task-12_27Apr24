import openpyxl
from datetime import datetime


# Method to open Excel file from specified path and read data from it.
def read_from_excel(xl_filepath):
    data = []
    workbook = openpyxl.load_workbook(xl_filepath)  # Opens the Excel file
    sheet = workbook.active  # Get active sheet
    # Iterates over rows and returns an iterator to iterate over rows in worksheet
    # and returns cell values
    for row in sheet.iter_rows(min_row=2, values_only=True):
        test_id = row[0]
        username = row[1]
        password = row[2]
        # This if condition makes sure none of the values returned is none and then
        # appends it to the object data
        if test_id is not None and username is not None and password is not None:
            data.append((test_id, username, password))
    return data


# Method to open Excel file from specified path and write data into it
def write_to_excel(xl_filepath, test_data, test_result):
    workbook = openpyxl.load_workbook(xl_filepath)
    sheet = workbook.active
    # Iterates over rows and returns an iterator to iterate over rows in worksheet
    # and returns cell objects
    for row in sheet.iter_rows(min_row=2, max_col=3, values_only=False):
        test_id_cell = row[0]
        username_cell = row[1]
        password_cell = row[2]

        # Checks if the values in the current row of the Excel sheet match the corresponding values in the test_data tuple.
        if test_id_cell.value == test_data[0] and username_cell.value == test_data[1] and password_cell.value == test_data[2]:
            # Write the Date and Time in separate columns
            # returns the reference of column 4 in same row as test_id_cell
            date_cell = sheet.cell(row=test_id_cell.row, column=4)
            time_cell = sheet.cell(row=test_id_cell.row, column=5)

            # Stores the date value into the cell
            # Stores the time value into the 5th column of same row as test_id_cell
            date_cell.value = datetime.now().date()
            time_cell.value = datetime.now().time().strftime("%H:%M:%S")

            # Write the Test Result in the 7th column
            result_cell = sheet.cell(row=test_id_cell.row, column=7)
            result_cell.value = test_result
    # Saves the workbook with the changes specified.
    workbook.save(xl_filepath)
